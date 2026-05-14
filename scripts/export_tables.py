"""Export the three weather-word spreadsheets to a single JSON file
the website can load.

Usage:
    python scripts/export_tables.py
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
OUTPUT = ROOT / "data" / "weather_words.json"

# Map sheets to (category, subcategory, iwi)
SHEET_MAP = {
    # Classification tables
    "Hauraki Catastrophe (7.1)": ("Classifications", "Catastrophe", "Hauraki"),
    "Apanui Catastrophe (8.5)":  ("Classifications", "Catastrophe", "Te Whānau-a-Apanui"),
    "Ngāi Tahu Catastrophe (9.2)": ("Classifications", "Catastrophe", "Ngāi Tahu"),
    "Hauraki Rain (7.4)":        ("Classifications", "Rain",       "Hauraki"),
    "Apanui Rain (8.9)":         ("Classifications", "Rain",       "Te Whānau-a-Apanui"),
    "Ngāi Tahu Rain (9.5)":      ("Classifications", "Rain",       "Ngāi Tahu"),
    "Hauraki Wind (7.5)":        ("Classifications", "Wind",       "Hauraki"),
    "Apanui Wind (8.10)":        ("Classifications", "Wind",       "Te Whānau-a-Apanui"),
    "Ngāi Tahu Wind (9.6)":      ("Classifications", "Wind",       "Ngāi Tahu"),
    "Hauraki Cloud (7.6)":       ("Classifications", "Cloud",      "Hauraki"),
    "Apanui Cloud (8.11)":       ("Classifications", "Cloud",      "Te Whānau-a-Apanui"),
    "Ngāi Tahu Cloud (9.7)":     ("Classifications", "Cloud",      "Ngāi Tahu"),

    # Indicators
    "Hauraki Weather (7.7)":     ("Indicators", "Weather",  "Hauraki"),
    "Apanui Weather (8.12)":     ("Indicators", "Weather",  "Te Whānau-a-Apanui"),
    "Ngāi Tahu Weather (9.8)":   ("Indicators", "Weather",  "Ngāi Tahu"),
    "Hauraki Climate (7.8)":     ("Indicators", "Climate",  "Hauraki"),
    "Apanui Climate (8.13)":     ("Indicators", "Climate",  "Te Whānau-a-Apanui"),
    "Ngāi Tahu Climate (9.9)":   ("Indicators", "Climate",  "Ngāi Tahu"),
    "Apanui Seasonal (8.14)":    ("Indicators", "Seasonal", "Te Whānau-a-Apanui"),
    "Ngāi Tahu Seasonal (9.10)": ("Indicators", "Seasonal", "Ngāi Tahu"),
}


def clean(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    # Normalize whitespace
    s = re.sub(r"\s+", " ", s)
    return s


def parse_sheet(ws, headers: list[str]) -> list[dict]:
    """Parse a sheet into rows of {header: value}."""
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        entry = {}
        for h, v in zip(headers, row):
            entry[h] = clean(v)
        # Skip empty rows
        if any(v for v in entry.values()):
            rows.append(entry)
    return rows


def main() -> None:
    data: dict[str, Any] = {
        "categories": {
            "Classifications": {},
            "Indicators": {},
        },
        "glossary": [],
    }

    # Classifications + Indicators
    for filename in ["classification_tables.xlsx", "weather_indicators.xlsx"]:
        path = ROOT / filename
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            if sheet_name not in SHEET_MAP:
                print(f"  Skipping unmapped sheet: {sheet_name}")
                continue
            category, subcategory, iwi = SHEET_MAP[sheet_name]
            ws = wb[sheet_name]

            # First row is headers
            headers = [clean(c) for c in next(ws.iter_rows(values_only=True))]
            rows = parse_sheet(ws, headers)

            # Strip the spreadsheet section number from the sheet name
            display = re.sub(r"\s*\(\d+\.\d+\)\s*$", "", sheet_name)

            data["categories"][category].setdefault(subcategory, {})[iwi] = {
                "sheet": display,
                "headers": headers,
                "rows": rows,
            }
            print(f"  {category} / {subcategory} / {iwi}: {len(rows)} rows")

    # Glossary
    path = ROOT / "glossary_he_kuputaka.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["He Kuputaka - Glossary"]
    headers = [clean(c) for c in next(ws.iter_rows(values_only=True))]
    glossary = parse_sheet(ws, headers)
    # Sort alphabetically by term
    glossary.sort(key=lambda r: (r.get(headers[0]) or "").lower())
    data["glossary"] = glossary
    print(f"  Glossary: {len(glossary)} terms")

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(data, ensure_ascii=False, indent=2))
    print(f"\nWrote {OUTPUT} ({OUTPUT.stat().st_size // 1024} KB)")


if __name__ == "__main__":
    main()